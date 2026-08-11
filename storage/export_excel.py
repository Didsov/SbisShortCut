import re
import os
import sqlite3
import tempfile
import zipfile
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXPORT_DIR = PROJECT_ROOT / "data" / "exports"
EXCEL_MAX_TEXT_LENGTH = 32767
ILLEGAL_XML_CHARACTERS = re.compile(
    "[\x00-\x08\x0B\x0C\x0E-\x1F]"
)


ALL_KKT_COLUMNS = (
    ("source_client_inn", "Исходный ИНН CRM"),
    ("source_client_kpp", "Исходный КПП CRM"),
    ("source_contractor_id", "Исходный ContractorId"),
    ("account_id", "AccountId"),
    ("account_name", "Аккаунт"),
    ("inn", "ИНН владельца"),
    ("kpp", "КПП владельца"),
    ("owner_name", "Название владельца"),
    ("kkt_contractor_id", "ContractorId владельца"),
    ("reg_number", "РНМ"),
    ("manufacturer_number", "ЗН ККТ"),
    ("fn_number", "ЗН ФН"),
    ("kkt_name", "Название ККТ"),
    ("model", "Модель"),
    ("address", "Адрес"),
    ("active", "Действующая"),
    ("status", "Статус"),
    ("fn_end_date", "Срок ФН"),
    ("ofd_end_date", "Срок ОФД"),
    ("parsed_at", "Добавлена"),
    ("updated_at", "Обновлена"),
)

REPLACEMENT_COLUMNS = (
    ("fn_end_date", "Срок ФН"),
    ("inn", "ИНН владельца"),
    ("owner_name", "Название владельца"),
    ("reg_number", "РНМ"),
    ("fn_number", "ЗН ФН"),
    ("account_id", "AccountId"),
    ("account_name", "Аккаунт"),
    ("model", "Модель"),
    ("address", "Адрес"),
)

IP_COLUMNS = (
    ("inn", "ИНН ИП"),
    ("ip_name", "ИП / аккаунт"),
    ("kpp", "КПП"),
    ("kkt_contractor_id", "ContractorId"),
    ("account_count", "Аккаунтов"),
    ("kkt_count", "Количество ККТ"),
    ("nearest_fn", "Ближайший срок ФН"),
    ("latest_fn", "Последний срок ФН"),
)

KKT_LOOKUP_COLUMNS = (
    ("owner_inn", "ИНН владельца"),
    ("owner_name", "ФИО владельца"),
    ("model", "Модель кассы"),
    ("reg_number", "РНМ"),
    ("manufacturer_number", "Заводской номер"),
    ("fn_end_date", "Срок ФН"),
    ("ofd_end_date", "Срок ОФД"),
)

DATABASE_KKT_COLUMNS = (
    ("inn", "ИНН"),
    ("owner_name", "Название организации"),
    ("address", "Адрес"),
    ("model", "Модель кассы"),
    ("reg_number", "Рег. номер"),
    ("manufacturer_number", "Заводской номер ККТ"),
    ("fn_number", "Заводской номер ФН"),
    ("fn_end_date", "Срок ФН"),
    ("ofd_end_date", "Срок ОФД"),
    ("owner_department_full_name", "Подразделение владельца"),
    ("ffd_version", "Версия ФФД"),
)


@dataclass(frozen=True)
class ExcelExportResult:
    database_path: Path
    output_path: Path
    mode: str
    row_count: int
    cutoff_date: str


def subtract_months(value: date, months: int) -> date:
    target_month = value.month - months
    target_year = value.year

    while target_month <= 0:
        target_month += 12
        target_year -= 1

    days = [
        31,
        29 if target_year % 400 == 0 or (
            target_year % 4 == 0 and target_year % 100 != 0
        ) else 28,
        31, 30, 31, 30, 31, 31, 30, 31, 30, 31,
    ]
    return date(target_year, target_month, min(value.day, days[target_month - 1]))


def sanitize_excel_value(value):
    if not isinstance(value, str):
        return value

    value = ILLEGAL_XML_CHARACTERS.sub(" ", value)

    if len(value) > EXCEL_MAX_TEXT_LENGTH:
        value = value[:EXCEL_MAX_TEXT_LENGTH]

    if value.startswith(("=", "+", "-", "@")):
        value = "'" + value

    return value


# SQLite stores these fields as TEXT.  Convert them before writing so Excel
# receives a serial date value rather than a string that only looks like one.
DATE_FIELDS = frozenset({"fn_end_date", "ofd_end_date", "parsed_at", "updated_at", "nearest_fn", "latest_fn"})


def excel_value(field: str, value):
    value = sanitize_excel_value(value)
    if field not in DATE_FIELDS or not isinstance(value, str):
        return value

    text = value.strip()
    if not text:
        return value

    try:
        # Preserve date-only values as date objects; Excel will display them
        # without an artificial midnight time.
        if len(text) == 10:
            return date.fromisoformat(text)
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        # Keep unusual source formats visible instead of dropping the value.
        return value


def readonly_connection(database_path: Path) -> sqlite3.Connection:
    database_path = Path(database_path).expanduser().resolve()

    if not database_path.is_file():
        raise FileNotFoundError(f"База не найдена: {database_path}")

    connection = sqlite3.connect(
        database_path.as_uri() + "?mode=ro",
        uri=True,
        timeout=30,
    )
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA query_only = ON")
    connection.execute("PRAGMA busy_timeout = 30000")
    return connection


def validate_kkt_table(connection: sqlite3.Connection) -> None:
    required = {
        field
        for field, _ in ALL_KKT_COLUMNS
        if field != "owner_name"
    }
    required.add("organization")
    actual = {
        row["name"]
        for row in connection.execute("PRAGMA table_info(kkt)")
    }
    missing = required - actual

    if missing:
        raise ValueError(
            "Таблица kkt отсутствует или не содержит поля: "
            + ", ".join(sorted(missing))
        )


def owner_name_sql(*fallbacks: str) -> str:
    """Название владельца из карточки контрагента, а не точки ККТ."""
    fallback_sql = ",\n                ".join(fallbacks)
    return f"""
        COALESCE(
            (
                SELECT NULLIF(TRIM(contractor.name), '')
                FROM contractors AS contractor
                WHERE contractor.inn = kkt.inn
                LIMIT 1
            ),
            (
                SELECT NULLIF(TRIM(client.name), '')
                FROM clients AS client
                WHERE client.inn = kkt.inn
                ORDER BY client.id DESC
                LIMIT 1
            ),
            {fallback_sql},
            '—'
        )
    """


def report_fields_sql(columns: tuple[tuple[str, str], ...]) -> str:
    fields = []

    for field, _ in columns:
        if field == "owner_name":
            fields.append(
                owner_name_sql(
                    "NULLIF(TRIM(kkt.organization), '')",
                ) + " AS owner_name"
            )
        else:
            fields.append(f"kkt.{field}")

    return ", ".join(fields)


def build_report_query(
    *,
    mode: str,
    today: date,
    days: int,
    ip_sort: str,
) -> tuple[str, tuple, tuple[tuple[str, str], ...], str, str]:
    cutoff = subtract_months(today, 3).isoformat()
    fresh_fn_filter = """
        fn_end_date IS NULL
        OR TRIM(fn_end_date) = ''
        OR DATE(fn_end_date) IS NULL
        OR DATE(fn_end_date) >= DATE(?)
    """

    if mode == "database":
        fields = f"""
            kkt.inn,
            {owner_name_sql("NULLIF(TRIM(kkt.organization), '')")} AS owner_name,
            kkt.address,
            kkt.model,
            kkt.reg_number,
            kkt.manufacturer_number,
            kkt.fn_number,
            kkt.fn_end_date,
            kkt.ofd_end_date,
            CASE
                WHEN JSON_VALID(kkt.raw_json)
                THEN JSON_EXTRACT(
                    kkt.raw_json,
                    '$.detail.Owner.DepartmentFullName'
                )
            END AS owner_department_full_name,
            CASE
                WHEN JSON_VALID(kkt.raw_json)
                THEN JSON_EXTRACT(
                    kkt.raw_json,
                    '$.detail.ffd_version'
                )
            END AS ffd_version
        """
        return (
            f"""
            SELECT {fields}
            FROM kkt AS kkt
            ORDER BY
                kkt.inn,
                owner_name COLLATE NOCASE,
                DATE(kkt.fn_end_date),
                kkt.reg_number
            """,
            (),
            DATABASE_KKT_COLUMNS,
            "Вся база ККТ",
            "не применяется",
        )

    if mode == "all":
        fields = report_fields_sql(ALL_KKT_COLUMNS)
        return (
            f"""
            SELECT {fields}
            FROM kkt AS kkt
            WHERE {fresh_fn_filter}
            ORDER BY
                kkt.inn,
                owner_name COLLATE NOCASE,
                DATE(kkt.fn_end_date),
                kkt.reg_number
            """,
            (cutoff,),
            ALL_KKT_COLUMNS,
            "ККТ",
            cutoff,
        )

    if mode == "replacements":
        date_from = (today - timedelta(days=7)).isoformat()
        date_to = (today + timedelta(days=max(0, days))).isoformat()
        fields = report_fields_sql(REPLACEMENT_COLUMNS)
        return (
            f"""
            SELECT {fields}
            FROM kkt AS kkt
            WHERE kkt.fn_end_date IS NOT NULL
              AND TRIM(kkt.fn_end_date) <> ''
              AND DATE(kkt.fn_end_date) BETWEEN DATE(?) AND DATE(?)
              AND ({fresh_fn_filter})
            ORDER BY
                DATE(kkt.fn_end_date),
                kkt.inn,
                owner_name COLLATE NOCASE,
                kkt.reg_number
            """,
            (date_from, date_to, cutoff),
            REPLACEMENT_COLUMNS,
            "Замена ФН",
            cutoff,
        )

    if mode == "ip":
        order_by = {
            "count": "kkt_count DESC, ip_name COLLATE NOCASE, inn",
            "name": "ip_name COLLATE NOCASE, inn",
            "fn": "DATE(nearest_fn), ip_name COLLATE NOCASE, inn",
        }.get(ip_sort)

        if order_by is None:
            raise ValueError(f"Неизвестная сортировка ИП: {ip_sort}")

        return (
            f"""
            SELECT
                kkt.inn,
                {owner_name_sql(
                    "MAX(NULLIF(TRIM(kkt.organization), ''))",
                    "MAX(NULLIF(TRIM(kkt.account_name), ''))",
                )} AS ip_name,
                MAX(kkt.kpp) AS kpp,
                MAX(kkt.kkt_contractor_id) AS kkt_contractor_id,
                COUNT(DISTINCT kkt.account_id) AS account_count,
                COUNT(DISTINCT kkt.reg_number) AS kkt_count,
                MIN(kkt.fn_end_date) AS nearest_fn,
                MAX(kkt.fn_end_date) AS latest_fn
            FROM kkt AS kkt
            WHERE kkt.inn IS NOT NULL
              AND LENGTH(TRIM(kkt.inn)) = 12
              AND ({fresh_fn_filter})
            GROUP BY kkt.inn
            ORDER BY {order_by}
            """,
            (cutoff,),
            IP_COLUMNS,
            "Сводка по ИП",
            cutoff,
        )

    raise ValueError(f"Неизвестный режим экспорта: {mode}")


def default_output_path(database_path: Path, mode: str) -> Path:
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f")
    return DEFAULT_EXPORT_DIR / (
        f"{Path(database_path).stem}_{mode}_{timestamp}.xlsx"
    )


def save_workbook_atomic(workbook: Workbook, output: Path) -> None:
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output.stem}_",
        suffix=".xlsx",
        dir=output.parent,
    )
    os.close(descriptor)
    Path(temporary_name).unlink(missing_ok=True)

    try:
        workbook.save(temporary_name)

        if not zipfile.is_zipfile(temporary_name):
            raise RuntimeError("Созданный файл не является корректным XLSX")

        Path(temporary_name).replace(output)
    finally:
        Path(temporary_name).unlink(missing_ok=True)


def export_kkt_lookup_excel(
    *,
    cash_registers,
    output_path: str | Path,
) -> int:
    """Экспортирует полный результат поиска ККТ по одному ИНН."""
    output = Path(output_path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ККТ по ИНН"
    worksheet.freeze_panes = "A2"
    worksheet.append([title for _, title in KKT_LOOKUP_COLUMNS])
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = [len(title) for _, title in KKT_LOOKUP_COLUMNS]
    row_count = 0

    for item in cash_registers:
        values = [
            excel_value(field, getattr(item, field, None))
            for field, _ in KKT_LOOKUP_COLUMNS
        ]
        worksheet.append(values)
        row_count += 1

        for index, (field, _) in enumerate(
            KKT_LOOKUP_COLUMNS,
            start=1,
        ):
            value = values[index - 1]

            if field in DATE_FIELDS and isinstance(value, (date, datetime)):
                worksheet.cell(
                    row=row_count + 1,
                    column=index,
                ).number_format = (
                    "yyyy-mm-dd"
                    if isinstance(value, date)
                    and not isinstance(value, datetime)
                    else "yyyy-mm-dd hh:mm:ss"
                )

        for index, value in enumerate(values):
            if value is not None:
                widths[index] = min(
                    50,
                    max(widths[index], len(str(value))),
                )

    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(KKT_LOOKUP_COLUMNS))}{row_count + 1}"
    )

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            width + 2,
            50,
        )

    save_workbook_atomic(workbook, output)
    return row_count


def export_kkt_excel(
    *,
    database_path: str | Path,
    mode: str = "all",
    output_path: str | Path | None = None,
    days: int = 30,
    ip_sort: str = "count",
    today: date | None = None,
) -> ExcelExportResult:
    database_path = Path(database_path).expanduser().resolve()
    report_date = today or date.today()
    query, params, columns, sheet_title, cutoff = build_report_query(
        mode=mode,
        today=report_date,
        days=days,
        ip_sort=ip_sort,
    )
    output = (
        Path(output_path).expanduser().resolve()
        if output_path is not None
        else default_output_path(database_path, mode).resolve()
    )
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31]
    worksheet.freeze_panes = "A2"
    worksheet.append([title for _, title in columns])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = [len(title) for _, title in columns]
    row_count = 0

    with closing(readonly_connection(database_path)) as connection:
        validate_kkt_table(connection)
        cursor = connection.execute(query, params)

        for row in cursor:
            values = [excel_value(field, row[field]) for field, _ in columns]
            worksheet.append(values)
            row_count += 1

            for index, (field, _) in enumerate(columns, start=1):
                if field in DATE_FIELDS and isinstance(values[index - 1], (date, datetime)):
                    worksheet.cell(row=row_count + 1, column=index).number_format = (
                        "yyyy-mm-dd" if isinstance(values[index - 1], date)
                        and not isinstance(values[index - 1], datetime)
                        else "yyyy-mm-dd hh:mm:ss"
                    )

            for index, value in enumerate(values):
                if value is not None:
                    widths[index] = min(
                        50,
                        max(widths[index], len(str(value))),
                    )

    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{row_count + 1}"

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            width + 2,
            50,
        )

    save_workbook_atomic(workbook, output)

    return ExcelExportResult(
        database_path=database_path,
        output_path=output,
        mode=mode,
        row_count=row_count,
        cutoff_date=cutoff,
    )
